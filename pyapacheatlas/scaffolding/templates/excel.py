from openpyxl import Workbook
from string import ascii_uppercase

COLUMN_TEMPLATE = [
    "Target Table", "Target Column", "Target Classifications",
    "Source Table", "Source Column", "Source Classifications",
    "Transformation"
]
TABLE_TEMPLATE = [
    "Target Table", "Target Type", "Target Classifications",
    "Source Table", "Source Type", "Source Classifications",
    "Process Name", "Process Type"
]
ENTITYDEF_TEMPLATE = [
    "Entity TypeName", "name", "description",
    "isOptional", "isUnique", "defaultValue",
    "typeName", "displayName", "valuesMinCount",
    "valuesMaxCount", "cardinality", "includeInNotification",
    "indexType", "isIndexable"
]

BULKENTITY_TEMPLATE = [
    "typeName", "name", "qualifiedName", "classifications"
]

def _update_sheet_headers(headers, worksheet):
    """
    For the given worksheet, make the first row equal to the list passed
    in as the headers.

    :param list headers: A list of column headers to use for this sheet.
    :param worksheet:
        The worksheet you are updating.
    :type worksheet:
        :class:`~openpyxl.worksheet.worksheet.Worksheet`
    """
    for idx, val in enumerate(headers):
        # TODO: Not the best way once we get past 26 columns in the template
        active_column = ascii_uppercase[idx]
        active_value = headers[idx]
        active_cell = "{}1".format(active_column)
        worksheet[active_cell] = active_value
        worksheet.column_dimensions[active_column].width = len(active_value)


def excel_template(filepath):
    """
    Generate an Excel template file and write it out to the given filepath.

    :param str filepath: The file path to store an XLSX file with the
        template Tables and Columns sheets.
    """
    wb = Workbook()
    columnsSheet = wb.active
    columnsSheet.title = "ColumnsLineage"
    tablesSheet = wb.create_sheet("TablesLineage")
    entityDefsSheet = wb.create_sheet("EntityDefs")
    bulkEntitiesSheet = wb.create_sheet("BulkEntities")

    _update_sheet_headers(COLUMN_TEMPLATE, columnsSheet)
    _update_sheet_headers(TABLE_TEMPLATE, tablesSheet)
    _update_sheet_headers(ENTITYDEF_TEMPLATE, entityDefsSheet)
    _update_sheet_headers(BULKENTITY_TEMPLATE, bulkEntitiesSheet)

    wb.save(filepath)
    wb.close()
