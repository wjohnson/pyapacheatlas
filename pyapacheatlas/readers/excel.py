from string import ascii_uppercase

from openpyxl import Workbook
from openpyxl import load_workbook

from .reader import Reader, ReaderConfiguration


class ExcelConfiguration(ReaderConfiguration):
    """
    A configuration utility to understand how your Excel file is structured.

    You must have a "Columns" and "Tables" sheet.  The name is configurable
    with the column_sheet and table_sheet properties.

    The Columns sheet must contain a "Source/Target" Column and Table header.
    Optionally, a Classifications column can be provided for each
    Source/Target.

    The Tables sheet must contain a "Source/Target" Table and Type along with
    a Process Name and Process Type.  The Process is related to the mechanism
    by which source becomes the target (e.g. a Stored Procedure or Query).

    :param str bulkEntity_sheet: Defaults to "BulkEntities"
    :param str updateLineage_sheet: Defaults to "UpdateLineage"
    :param str columnMapping_sheet: Defaults to "ColumnMapping"
    :param str entityDef_sheet: Defaults to "EntityDefs"
    :param str classificationDef_sheet: Defaults to "ClassificationDefs"
    :param str table_sheet: Defaults to "TablesLineage"
    :param str column_sheet: Defaults to "FineGrainColumnLineage"
    :param str source_prefix:
        Defaults to "Source" and represents the prefix of the columns
        in Excel to be considered related to the source table or column.
    :param str target_prefix:
        Defaults to "Target" and represents the prefix of the columns
        in Excel to be considered related to the target table or column.
    :param str process_prefix:
        Defaults to "Process" and represents the prefix of the columns
        in Excel to be considered related to the table process.
    :param str column_transformation_name:
        Defaults to "transformation" and identifies the column that
        represents the transformation for a specific column.
    """

    def __init__(self, column_sheet="FineGrainColumnLineage",
                 table_sheet="TablesLineage",
                 entityDef_sheet="EntityDefs", bulkEntity_sheet="BulkEntities",
                 classificationDef_sheet="ClassificationDefs",
                 updateLineage_sheet="UpdateLineage",
                 columnMapping_sheet="ColumnMapping",
                 **kwargs):
        super().__init__(**kwargs)
        # Required attributes:
        # qualifiedName, column, transformation, table
        self.column_sheet = column_sheet
        self.table_sheet = table_sheet
        self.entityDef_sheet = entityDef_sheet
        self.classificationDef_sheet = classificationDef_sheet
        self.bulkEntity_sheet = bulkEntity_sheet
        self.updateLineage_sheet = updateLineage_sheet
        self.columnMapping_sheet = columnMapping_sheet


class ExcelReader(Reader):
    """
    Read in Excel files that follow the excel template tab structure.
    Expects an :class:`~pyapacheatlas.readers.excel.ExcelConfiguration` object
    to determine the naming conventions of tabs and headers.
    """
    @staticmethod
    def _parse_spreadsheet(worksheet):
        """
        Standardizes the excel worksheet into a json format.

        :param openpyxl.workbook.Workbook worksheet:
            A worksheet class from openpyxl.
        :return:
            The standardized version of the excel spreadsheet in json form.
        :rtype: list(dict(str,str))
        """
        # Standardize the column header
        # TODO: Allow for skip lines
        column_headers = list(
            zip(
                range(0, len(worksheet[1])),
                [str(c.value).strip() for c in worksheet[1]]
            )
        )

        output = []
        for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row,
                min_col=1, max_col=worksheet.max_column):
            output.append(
                {k: row[idx].value for idx, k in column_headers})

        return output

    def parse_bulk_entities(self, filepath, contacts_func=None):
        """
        Generate a set of entities from an excel template file.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :param function contacts_func:
            For Azure Purview, a function to be called on each value
            when you pass in an experts or owners header to json_rows.
            Leaving it as None will return the exact value passed in
            to the experts and owners section.
            It has a built in cache that will prevent redundant calls
            to your function.

        :return: An AtlasTypeDef with entityDefs for the provided rows.
        :rtype: dict(str, list(dict))
        """
        wb = load_workbook(filepath)

        sheetIsNotPresent = self.config.bulkEntity_sheet not in wb.sheetnames
        if self.config.bulkEntity_sheet and sheetIsNotPresent:
            raise KeyError("The sheet {} was not found".format(
                self.config.entityDef_sheet))

        output = dict()

        if self.config.bulkEntity_sheet:
            bulkEntity_sheet = wb[self.config.bulkEntity_sheet]
            json_bulkEntities = ExcelReader._parse_spreadsheet(
                bulkEntity_sheet)
            bulkEntities_generated = super().parse_bulk_entities(
                json_bulkEntities, contacts_func)
            output.update(bulkEntities_generated)

        wb.close()

        return output

    def parse_entity_defs(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the type def tab(s) into a set of entity defs that can be
        uploaded.

        Currently, only entityDefs are supported.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return: An AtlasTypeDef with entityDefs for the provided rows.
        :rtype: dict(str, list(dict))
        """
        wb = load_workbook(filepath)
        # A user may omit the entityDef_sheet by providing the config with None
        sheetIsNotPresent = self.config.entityDef_sheet not in wb.sheetnames
        if self.config.entityDef_sheet and sheetIsNotPresent:
            raise KeyError("The sheet {} was not found".format(
                self.config.entityDef_sheet))

        output = dict()

        # Getting entityDefinitions if the user provided a name of the sheet
        if self.config.entityDef_sheet:
            entityDef_sheet = wb[self.config.entityDef_sheet]
            json_entitydefs = ExcelReader._parse_spreadsheet(entityDef_sheet)
            entityDefs_generated = super().parse_entity_defs(json_entitydefs)
            output.update(entityDefs_generated)

        wb.close()
        # TODO: Add in classificationDefs and relationshipDefs
        return output

    def parse_finegrain_column_lineage(self, filepath, atlas_entities, atlas_typedefs, use_column_mapping=False):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the columns into column lineages.
        Requires that the relationship attributes are already defined in the
        provided atlas type defs.

        Infers column type from the target table type and an assumed "columns"
        relationship attribute on the table type.

        Infers the column lineage process based on the provided table process
        (provided in the template's table excel sheet).  Looks for the first
        relationship type def with an endDef2 of `columnLineages`.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :param list() atlas_entities:
            A list of AtlasEntity objects representing
        :param dict(str,list(dict)) atlas_typedefs:
            The results of requesting all type defs from Apache Atlas,
            including entityDefs, relationshipDefs, etc.  relationshipDefs
            are the only values used.
        :param bool use_column_mapping:
            Should the table processes include the columnMappings attribute
            that represents Column Lineage in Azure Data Catalog.
            Defaults to False.
        :return:
            A list of Atlas Entities representing the spreadsheet's
            inputs as their json dicts.
        :rtype: list(:class:`~pyapacheatlas.core.entity.AtlasEntity`)
        """

        wb = load_workbook(filepath)

        if self.config.column_sheet not in wb.sheetnames:
            raise KeyError("The sheet {} was not found".format(
                self.config.column_sheet))

        # Getting column entities
        column_sheet = wb[self.config.column_sheet]
        json_columns = ExcelReader._parse_spreadsheet(column_sheet)

        entities = super().parse_finegrain_column_lineage(
            json_columns,
            atlas_entities,
            atlas_typedefs,
            use_column_mapping=use_column_mapping
        )

        wb.close()

        return entities

    def parse_table_lineage(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the tables and processes table lineages.
        Requires that the relationship attributes are already defined in the
        provided atlas type defs.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return:
            A list of Atlas Entities representing the spreadsheet's inputs
            as their json dicts.
        :rtype: list(:class:`~pyapacheatlas.core.entity.AtlasEntity`)
        """

        wb = load_workbook(filepath)

        entities = []

        if self.config.table_sheet not in wb.sheetnames:
            raise KeyError("The sheet {} was not found".format(
                self.config.table_sheet))

        # Getting table entities
        table_sheet = wb[self.config.table_sheet]
        json_sheet = ExcelReader._parse_spreadsheet(table_sheet)
        entities = super().parse_table_lineage(json_sheet)

        wb.close()

        return entities

    def parse_table_finegrain_column_lineages(self, filepath, atlas_typedefs, use_column_mapping=False):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the tables, processes, and columns into table and column
        lineages. Requires that the relationship attributes are already
        defined in the provided atlas type defs.

        Infers column type from the target table type and an assumed "columns"
        relationship attribute on the table type.

        Infers the column lineage process based on the provided table process
        (provided in the template's table excel sheet).  Looks for the first
        relationship type def with an endDef2 of `columnLineages`.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :param dict(str,list(dict)) atlas_typedefs:
            The results of requesting all type defs from Apache Atlas,
            including entityDefs, relationshipDefs, etc.  relationshipDefs are
            the only values used.
        :param bool use_column_mapping:
            Should the table processes include the columnMappings attribute
            that represents Column Lineage in Azure Data Catalog.
            Defaults to False.
        :return:
            A list of Atlas Entities representing the spreadsheet's inputs
            as their json dicts.
        :rtype: list(dict)
        """
        entities = []

        table_entities = self.parse_table_lineage(filepath)
        entities.extend(table_entities)

        # Modifies table_entities if use_column_mapping is True
        column_entities = self.parse_finegrain_column_lineage(
            filepath,
            table_entities,
            atlas_typedefs,
            use_column_mapping
        )
        entities.extend(column_entities)

        output = [e.to_json() for e in entities]

        return output

    def parse_update_lineage(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the (default) UpdateLineage table into existing process entities. 

        Assumes these process entities and any referenced entity exists.

        Leave the qualifiedName cell blank on source or target to leave the
        existing input or output (respectively) unchanged.

        Use 'N/A' in the qualifiedName on source or target to 'destroy' the
        existing input or output and overwrite with an empty list.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return:
            A list of Atlas Process entities representing the spreadsheet's
            contents.
        :rtype: list(dict)
        """
        wb = load_workbook(filepath)

        entities = []

        if self.config.updateLineage_sheet not in wb.sheetnames:
            raise KeyError("The sheet {} was not found".format(
                self.config.updateLineage_sheet))

        # Getting table entities
        updateLineage_sheet = wb[self.config.updateLineage_sheet]
        json_sheet = ExcelReader._parse_spreadsheet(updateLineage_sheet)
        entities = super().parse_update_lineage(json_sheet)

        wb.close()

        return entities

    def parse_column_mapping(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the (default) ColumnMapping tab into existing process entities. 

        Assumes these process entities and any referenced entity exists.
        This will not update the inputs and outputs, it will update name
        and columnMapping fields.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return:
            A list of Atlas Process entities representing the spreadsheet's
            contents.
        :rtype: list(dict)
        """
        wb = load_workbook(filepath)

        entities = []

        if self.config.columnMapping_sheet not in wb.sheetnames:
            raise KeyError("The sheet {} was not found".format(
                self.config.columnMapping_sheet))

        # Getting table entities
        columnMapping_sheet = wb[self.config.columnMapping_sheet]
        json_sheet = ExcelReader._parse_spreadsheet(columnMapping_sheet)
        entities = super().parse_column_mapping(json_sheet)

        wb.close()

        return entities

    def parse_update_lineage_with_mappings(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the (default) UpdateLineage and ColumnMapping tabs into existing process entities. 

        Assumes these process entities and any referenced entity exists.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return:
            A list of Atlas Process entities representing the spreadsheet's
            contents.
        :rtype: list(dict)
        """

        lineage = self.parse_update_lineage(filepath)
        mappings = self.parse_column_mapping(filepath)
        seen_qualifiedNames = {}
        for working_entity in lineage + mappings:
            qn = working_entity["attributes"]["qualifiedName"]
            if qn in seen_qualifiedNames:
                # If we have seen an entity before check if
                # the working entity contains a column mapping attribute
                # if it does update the existing entity
                if "columnMapping" in working_entity["attributes"]:
                    seen_qualifiedNames[qn]["attributes"]["columnMapping"] = working_entity["attributes"]["columnMapping"]
            else:
                # If we haven't seen it just add the entity to the list
                seen_qualifiedNames[qn] = working_entity

        return list(seen_qualifiedNames.values())

    def parse_classification_defs(self, filepath):
        """
        Read a given excel file that conforms to the excel atlas template and
        parse the (default) ClassificationDefs tab into classifications.

        :param str filepath:
            The xlsx file that contains your table and columns.
        :return:
            A  entities representing the spreadsheet's
            contents.
        :rtype: list(dict)
        """
        wb = load_workbook(filepath)
        # A user may omit the classificationDef_sheet by providing the config with None
        sheetIsNotPresent = self.config.classificationDef_sheet not in wb.sheetnames
        if self.config.classificationDef_sheet and sheetIsNotPresent:
            raise KeyError("The sheet {} was not found".format(
                self.config.classificationDef_sheet))

        output = dict()

        # Getting classificationDef if the user provided a name of the sheet
        if self.config.classificationDef_sheet:
            classificationDef_sheet = wb[self.config.classificationDef_sheet]
            json_classificationdefs = ExcelReader._parse_spreadsheet(
                classificationDef_sheet)
            classificationDefs_generated = super(
            ).parse_classification_defs(json_classificationdefs)
            output.update(classificationDefs_generated)

        wb.close()
        return output

    @staticmethod
    def _update_sheet_headers(headers, worksheet):
        """
        For the given worksheet, make the first row equal to the list passed
        in as the headers.

        :param list headers: A list of column headers to use for this sheet.
        :param worksheet:
            The worksheet you are updating.
        :type worksheet:
            :class:`~openpyxl.worksheet.worksheet.Worksheet`
        """
        for idx, val in enumerate(headers):
            # TODO: Not the best way once we get past
            # 26 columns in the template
            active_column = ascii_uppercase[idx]
            active_value = headers[idx]
            active_cell = "{}1".format(active_column)
            worksheet[active_cell] = active_value
            worksheet.column_dimensions[active_column].width = len(
                active_value)

    @staticmethod
    def _replace_header_prefix(headers, prefix_mapping):
        new_header = []
        for column in headers:
            first_token, *remainder = column.split(" ", 1)
            if first_token in prefix_mapping:
                new_header.append(
                    prefix_mapping[first_token]+' '+''.join(remainder))
            else:
                new_header.append(column)
        return new_header

    @staticmethod
    def make_template(filepath, **kwargs):
        """
        Generate an Excel template file and write it out to the given filepath.

        :param str filepath: The file path to store an XLSX file with the
            template Tables and Columns sheets.
        :param str bulkEntity_sheet: Defaults to "BulkEntities"
        :param str updateLineage_sheet: Defaults to "UpdateLineage"
        :param str columnMapping_sheet: Defaults to "ColumnMapping"
        :param str entityDef_sheet: Defaults to "EntityDefs"
        :param str classificationDef_sheet: Defaults to "ClassificationDefs"
        :param bool include_deprecated:
            Set to True if you want to include tabs that have been deprecated.
            For this release, it includes TablesLineage and
            FineGrainColumnLineage.
        :param str table_sheet: Defaults to "TablesLineage"
        :param str column_sheet: Defaults to "FineGrainColumnLineage"
        :param str source_prefix:
            Defaults to "Source" and represents the prefix of the columns
            in Excel to be considered related to the source table or column.
        :param str target_prefix:
            Defaults to "Target" and represents the prefix of the columns
            in Excel to be considered related to the target table or column.
        :param str process_prefix:
            Defaults to "Process" and represents the prefix of the columns
            in Excel to be considered related to the table process.
        :param str column_transformation_name:
            Defaults to "transformation" and identifies the column that
            represents the transformation for a specific column.
        """
        include_deprecated = kwargs.get("include_deprecated", False)
        wb = Workbook()
        bulkEntitiesSheet = wb.active
        bulkEntitiesSheet.title = kwargs.get(
            "bulkEntity_sheet", "BulkEntities")
        updateLineageSheet = wb.create_sheet(
            kwargs.get("updateLineage_sheet", "UpdateLineage"))
        columnMappingSheet = wb.create_sheet(
            kwargs.get("columnMapping_sheet", "ColumnMapping"))
        entityDefsSheet = wb.create_sheet(
            kwargs.get("entityDef_sheet", "EntityDefs"))
        classificationDefsSheet = wb.create_sheet(kwargs.get(
            "classificationDef_sheet", "ClassificationDefs"))
        if include_deprecated:
            tablesSheet = wb.create_sheet(
                kwargs.get("table_sheet", "TablesLineage"))
            columnsSheet = wb.create_sheet(kwargs.get(
                "column_sheet", "FineGrainColumnLineage"))

        # Supporting changing the default headers on select pages
        header_changes = {}
        if "source_prefix" in kwargs:
            header_changes["Source"] = kwargs["source_prefix"]
        if "target_prefix" in kwargs:
            header_changes["Target"] = kwargs["target_prefix"]
        if "process_prefix" in kwargs:
            header_changes["Process"] = kwargs["process_prefix"]
        if "column_transformation_name" in kwargs:
            header_changes["transformation"] = kwargs["column_transformation_name"]

        if header_changes:
            FineGrainColumnLineageHeaders = ExcelReader._replace_header_prefix(
                Reader.TEMPLATE_HEADERS["FineGrainColumnLineage"],
                header_changes
            )
            TablesLineageHeaders = ExcelReader._replace_header_prefix(
                Reader.TEMPLATE_HEADERS["TablesLineage"],
                header_changes
            )
            UpdateLineageHeaders = ExcelReader._replace_header_prefix(
                Reader.TEMPLATE_HEADERS["UpdateLineage"],
                header_changes
            )
            ColumnMappingHeaders = ExcelReader._replace_header_prefix(
                Reader.TEMPLATE_HEADERS["ColumnMapping"],
                header_changes
            )
        else:
            FineGrainColumnLineageHeaders = Reader.TEMPLATE_HEADERS["FineGrainColumnLineage"]
            TablesLineageHeaders = Reader.TEMPLATE_HEADERS["TablesLineage"]
            UpdateLineageHeaders = Reader.TEMPLATE_HEADERS["UpdateLineage"]
            ColumnMappingHeaders = Reader.TEMPLATE_HEADERS["ColumnMapping"]

        if include_deprecated:
            ExcelReader._update_sheet_headers(
                FineGrainColumnLineageHeaders, columnsSheet
            )
            ExcelReader._update_sheet_headers(
                TablesLineageHeaders, tablesSheet
            )
        ExcelReader._update_sheet_headers(
            Reader.TEMPLATE_HEADERS["EntityDefs"], entityDefsSheet
        )
        ExcelReader._update_sheet_headers(
            Reader.TEMPLATE_HEADERS["ClassificationDefs"], classificationDefsSheet
        )
        ExcelReader._update_sheet_headers(
            Reader.TEMPLATE_HEADERS["BulkEntities"], bulkEntitiesSheet
        )
        ExcelReader._update_sheet_headers(
            UpdateLineageHeaders, updateLineageSheet
        )
        ExcelReader._update_sheet_headers(
            ColumnMappingHeaders, columnMappingSheet
        )

        wb.save(filepath)
        wb.close()
