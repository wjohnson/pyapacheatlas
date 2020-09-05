import os

from openpyxl import load_workbook

from pyapacheatlas.scaffolding.templates import excel_template

def test_verify_template_sheets():
    # Setup
    temp_path = "./temp_verfiysheets.xlsx"
    excel_template(temp_path)

    # Expected
    expected_sheets = set(["ColumnsLineage", "TablesLineage", 
        "EntityDefs", "BulkEntities"
    ])

    wb = load_workbook(temp_path)
    difference = set(wb.sheetnames).symmetric_difference(expected_sheets)
    try:
        assert(len(difference) == 0)
    finally:
        wb.close()
        os.remove(temp_path)

