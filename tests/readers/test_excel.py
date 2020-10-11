
import json
import os

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


from pyapacheatlas.readers.excel import ExcelConfiguration, ExcelReader
from pyapacheatlas.scaffolding.column_lineage import column_lineage_scaffold

def test_verify_template_sheets():
    # Setup
    temp_path = "./temp_verfiysheets.xlsx"
    ExcelReader.make_template(temp_path)

    # Expected
    expected_sheets = set(["ColumnsLineage", "TablesLineage",
                           "EntityDefs", "BulkEntities",
                           "UpdateLineage"
                           ])

    wb = load_workbook(temp_path)
    difference = set(wb.sheetnames).symmetric_difference(expected_sheets)
    try:
        assert(len(difference) == 0)
    finally:
        wb.close()
        os.remove(temp_path)


def setup_workbook_custom_sheet(filepath, sheet_name, headers, json_rows):
    wb = Workbook()
    customSheet = wb.active
    customSheet.title = sheet_name
    ExcelReader._update_sheet_headers(headers, customSheet)
    row_counter = 0
    # TODO: Clear the column headers
    # Add the data to the sheet
    for row in customSheet.iter_rows(min_row=2, max_col=len(headers), max_row=len(json_rows)+1):
        for idx, cell in enumerate(row):
            cell.value = json_rows[row_counter][idx]
        row_counter += 1

    wb.save(filepath)
    wb.close()


def setup_workbook(filepath, sheet_name, max_col, json_rows):
    if not os.path.exists(filepath):
        ExcelReader.make_template(filepath)
    wb = load_workbook(filepath)
    active_sheet = wb[sheet_name]

    row_counter = 0
    # TODO: Make the max_col more dynamic
    for row in active_sheet.iter_rows(min_row=2, max_col=max_col, max_row=len(json_rows)+1):
        for idx, cell in enumerate(row):
            cell.value = json_rows[row_counter][idx]
        row_counter += 1

    wb.save(filepath)


def remove_workbook(filepath):
    if os.path.exists(filepath) and os.path.isfile(filepath):
        os.remove(filepath)


def test_excel_typeDefs_entityTypes():
    temp_filepath = "./temp_test_typeDefs_entityTYpes.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)
    max_cols = len(ExcelReader.TEMPLATE_HEADERS["BulkEntities"])
    # "Entity TypeName", "name", "description",
    # "isOptional", "isUnique", "defaultValue",
    # "typeName", "displayName", "valuesMinCount",
    # "valuesMaxCount", "cardinality", "includeInNotification",
    # "indexType", "isIndexable"
    json_rows = [
        ["demoType", "attrib1", "Some desc",
         "True", "False", None,
         "string", None, None,
         None, None, None,
         None, None
         ]
    ]
    setup_workbook(temp_filepath, "EntityDefs", max_cols, json_rows)

    results = reader.parse_entity_defs(temp_filepath)

    assert("entityDefs" in results)
    assert(len(results["entityDefs"]) == 1)
    assert(results["entityDefs"][0]["attributeDefs"][0]["name"] == "attrib1")

    remove_workbook(temp_filepath)


def test_excel_bulkEntities():
    temp_filepath = "./temp_test_excel_bulkEntities.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)
    max_cols = len(ExcelReader.TEMPLATE_HEADERS["BulkEntities"])
    # "typeName", "name",
    # "qualifiedName", "classifications"
    json_rows = [
        ["demoType", "entityNameABC",
         "qualifiedNameofEntityNameABC", None
         ],
        ["demoType", "entityNameGHI",
         "qualifiedNameofEntityNameGHI", None
         ]
    ]
    setup_workbook(temp_filepath, "BulkEntities", max_cols, json_rows)

    results = reader.parse_bulk_entities(temp_filepath)

    try:
        assert("entities" in results)
        assert(len(results["entities"]) == 2)
    finally:
        remove_workbook(temp_filepath)


def test_excel_bulkEntities_withClassifications():
    temp_filepath = "./temp_test_excel_bulkEntitiesWithClassifications.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)
    max_cols = len(ExcelReader.TEMPLATE_HEADERS["BulkEntities"])
    # "typeName", "name",
    # "qualifiedName", "classifications"
    json_rows = [
        ["demoType", "entityNameABC",
         "qualifiedNameofEntityNameABC", "PII"
         ],
        ["demoType", "entityNameGHI",
         "qualifiedNameofEntityNameGHI", "PII;CLASS2"
         ]
    ]

    setup_workbook(temp_filepath, "BulkEntities", max_cols, json_rows)

    results = reader.parse_bulk_entities(temp_filepath)

    try:
        assert("entities" in results)
        assert(len(results["entities"]) == 2)
        abc = results["entities"][0]
        ghi = results["entities"][1]

        assert(len(abc["classifications"]) == 1)
        assert(len(ghi["classifications"]) == 2)

        assert(abc["classifications"][0]["typeName"] == "PII")
        ghi_classification_types = set(
            [x["typeName"] for x in ghi["classifications"]]
        )
        assert(set(["PII", "CLASS2"]) == ghi_classification_types)
    finally:
        remove_workbook(temp_filepath)


def test_excel_bulkEntities_dynamicAttributes():
    temp_filepath = "./temp_test_excel_bulkEntitieswithAttributes.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)

    headers = ExcelReader.TEMPLATE_HEADERS["BulkEntities"] + \
        ["attrib1", "attrib2"]
    # "typeName", "name",
    # "qualifiedName", "classifications"
    # "attrib1", "attrib2"
    json_rows = [
        ["demoType", "entityNameABC",
         "qualifiedNameofEntityNameABC", None,
         None, "abc"
         ],
        ["demoType", "entityNameGHI",
         "qualifiedNameofEntityNameGHI", None,
         "ghi", "abc2"
         ]
    ]

    setup_workbook_custom_sheet(
        temp_filepath, "BulkEntities", headers, json_rows)

    results = reader.parse_bulk_entities(temp_filepath)

    try:
        assert("entities" in results)
        assert(len(results["entities"]) == 2)
        abc = results["entities"][0]
        ghi = results["entities"][1]

        assert("attrib1" not in abc["attributes"])
        assert("attrib2" in abc["attributes"])
        assert(abc["attributes"]["attrib2"] == "abc")

        assert("attrib1" in ghi["attributes"])
        assert("attrib2" in ghi["attributes"])
        assert(ghi["attributes"]["attrib2"] == "abc2")
        assert(ghi["attributes"]["attrib1"] == "ghi")

    finally:
        remove_workbook(temp_filepath)


def test_excel_table_lineage():
    temp_filepath = "./temp_test_excel_table_lineage.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)
    max_cols = len(ExcelReader.TEMPLATE_HEADERS["TablesLineage"])

    # "Target Table", "Target Type", "Target Classifications",
    # "Source Table", "Source Type", "Source Classifications",
    # "Process Name", "Process Type"

    json_rows = [
        ["table1", "demo_type", None,
         "table0", "demo_type2", None,
         "proc01", "proc_type"
         ]
    ]

    setup_workbook(temp_filepath, "TablesLineage", max_cols, json_rows)

    results = reader.parse_table_lineage(temp_filepath)

    try:
        assert(results[0].to_json(minimum=True) == {
            "typeName": "demo_type", "guid": -1001, "qualifiedName": "table1"})
        assert(results[1].to_json(minimum=True) == {
            "typeName": "demo_type2", "guid": -1002, "qualifiedName": "table0"})
        assert(results[2].to_json(minimum=True) == {
            "typeName": "proc_type", "guid": -1003, "qualifiedName": "proc01"})
    finally:
        remove_workbook(temp_filepath)


def test_excel_column_lineage():
    temp_filepath = "./temp_test_excel_column_lineage.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)
    max_cols_tl = len(ExcelReader.TEMPLATE_HEADERS["TablesLineage"])
    max_cols_cl = len(ExcelReader.TEMPLATE_HEADERS["ColumnsLineage"])

    # "Target Table", "Target Type", "Target Classifications",
    # "Source Table", "Source Type", "Source Classifications",
    # "Process Name", "Process Type"

    json_rows = [
        ["table1", "demo_table", None,
         "table0", "demo_table", None,
         "proc01", "demo_process"
         ]
    ]

    # "Target Table", "Target Column", "Target Classifications",
    # "Source Table", "Source Column", "Source Classifications",
    # "Transformation"
    json_rows_col = [
        ["table1", "t00", None,
         "table0", "t00", None,
         None],
        ["table1", "tcombo", None,
         "table0", "tA", None,
         None],
        ["table1", "tcombo", None,
         "table0", "tB", None,
         None],
    ]

    setup_workbook(temp_filepath, "TablesLineage", max_cols_tl, json_rows)
    setup_workbook(temp_filepath, "ColumnsLineage", max_cols_cl, json_rows_col)

    atlas_types = column_lineage_scaffold("demo")
    
    table_entities = reader.parse_table_lineage(temp_filepath)

    # For column mappings, table_entities do not contain columnMapping
    assert(all(["columnMapping" not in e.attributes for e in table_entities]))

    column_entities = reader.parse_column_lineage(temp_filepath, 
        table_entities,
        atlas_types, 
        use_column_mapping= True
    )

    try:
        table1 = None
        table0 = None
        proc01 = None
        t00 = None
        table1_t00 = None
        table0_t00 = None
        col_lineage_process = None
        table_lookup = {e.get_name():e for e in table_entities}
        column_lookup = {e.get_name():e for e in column_entities}
        
        # We have five columns (t00 > t00) + ((tA + tB) > tcombo) 
        # and two processes
        assert(len(column_entities) == 7)

        # Because of column mappings is TRUE, table entities are modified
        assert("columnMapping" in table_lookup["proc01"].attributes)
        resulting_col_map = json.loads(table_lookup["proc01"].attributes["columnMapping"])[0]
        expected_col_map = {
            "DatasetMapping":{"Source":"table0", "Sink":"table1"},
            "ColumnMapping":[
                {"Source":"t00","Sink":"t00"},
                {"Source":"tA","Sink":"tcombo"},
                {"Source":"tB","Sink":"tcombo"}
            ]
        }
        assert(resulting_col_map["DatasetMapping"] == expected_col_map["DatasetMapping"])
        assert(len(resulting_col_map["ColumnMapping"]) == 3)
        assert(resulting_col_map["ColumnMapping"][0] in expected_col_map["ColumnMapping"])
        assert(resulting_col_map["ColumnMapping"][1] in expected_col_map["ColumnMapping"])
        assert(resulting_col_map["ColumnMapping"][2] in expected_col_map["ColumnMapping"])
    finally:
        remove_workbook(temp_filepath)

def test_excel_update_lineage():
    temp_filepath = "./temp_test_excel_updateLineage.xlsx"
    ec = ExcelConfiguration()
    reader = ExcelReader(ec)

    headers = ExcelReader.TEMPLATE_HEADERS["UpdateLineage"]

    # Same as main test
    json_rows = [
        [
        "demo_table", "demotarget", "demo_table2", "demosource",
         "proc01", "procqual01", "Process2"
        ]
    ]

    setup_workbook_custom_sheet(
        temp_filepath, "UpdateLineage", headers, json_rows)

    results = reader.parse_update_lineage(temp_filepath)

    try:
        assert(len(results) == 1)
    finally:
        remove_workbook(temp_filepath)