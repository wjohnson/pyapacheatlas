import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient, AtlasEntity
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader


def fill_in_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    entityDef_sheet = wb[excel_config.entityDef_sheet]
    updateLineage_sheet = wb[excel_config.updateLineage_sheet]
    mapping_sheet = wb[excel_config.columnMapping_sheet]
    

    # Update Lineage Sheet SCHEMA
    # "Target typeName", "Target qualifiedName", "Source typeName",
    # "Source qualifiedName", "Process name", "Process qualifiedName",
    # "Process typeName"
    lineage_to_update = [
        ["azure_sql_table", "pyapacheatlas://demo_update_lineagemapping_output",
         "azure_sql_table", "pyapacheatlas://demo_update_lineagemapping_input",
         "demo_some_sql_process",
         "pyapacheatlas://process_update_lineage_custom_sql_process",
         "customProcessWithMapping"
         ]
    ]

    # Mapping SCHEMA
    # "Source qualifiedName", "Source column", "Target qualifiedName", 
    # "Target column", "Process qualifiedName", "Process typeName",
    # "Process name"
    mapping_to_update = [
        ["pyapacheatlas://demo_update_lineagemapping_input", "A1", "pyapacheatlas://demo_update_lineagemapping_output",
        "B1", "pyapacheatlas://process_update_lineage_custom_sql_process", "customProcessWithMapping",
        "demo_some_sql_process"
        ],
        ["pyapacheatlas://demo_update_lineagemapping_input", "A2", "pyapacheatlas://demo_update_lineagemapping_output",
        "BCombo", "pyapacheatlas://process_update_lineage_custom_sql_process", "customProcessWithMapping",
        "demo_some_sql_process"
        ],
        ["pyapacheatlas://demo_update_lineagemapping_input", "A3", "pyapacheatlas://demo_update_lineagemapping_output",
        "BCombo", "pyapacheatlas://process_update_lineage_custom_sql_process", "customProcessWithMapping",
        "demo_some_sql_process"
        ]
    ]

    # EntityDef SCHEMA
    # "Entity TypeName", "name", "description",
    # "isOptional", "isUnique", "defaultValue",
    # "typeName", "displayName", "valuesMinCount",
    # "valuesMaxCount", "cardinality", "includeInNotification",
    # "indexType", "isIndexable", Entity superTypes (as an added non-default)
    entitydef_to_update = [
        [
        "customProcessWithMapping", "columnMapping", "stringified json to support mappings in Purview UI",
        True, None, None,
        "string", None, None,
        None, None, None,
        None, None,
        "Process"
        ]
    ]

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in updateLineage_sheet.iter_rows(min_row=2, max_col=7,
                                             max_row=len(lineage_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = lineage_to_update[table_row_counter][idx]
        table_row_counter += 1
    
    # Update Column Mapping Sheet
    mapping_row_counter = 0
    for row in mapping_sheet.iter_rows(min_row=2, max_col=7,
                                             max_row=len(mapping_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = mapping_to_update[mapping_row_counter][idx]
        mapping_row_counter += 1

    # Update Entity Def Sheet
    entity_def_counter = 0
    for row in entityDef_sheet.iter_rows(min_row=2, max_col=15,
                                             max_row=len(entitydef_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = entitydef_to_update[entity_def_counter][idx]
        entity_def_counter += 1
    
    entityDef_sheet['O1'].value = 'Entity superTypes'

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file and
    generating a set of entities that would create or update a Process entity
    that links (applies lineage to) an input and output entity that already
    exists in your data catalog. In addition, it provides an example of using
    the Purview columnMapping feature. This sample uses built-in Azure types
    and will not work on Apache Atlas.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", ""),
        authentication=oauth
    )

    # SETUP: This is just setting up the excel file for you
    file_path = "./demo_update_lineage_with_mapping_upload.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # We are going to cheat here and create some entities before
    # we get to parsing the spreadsheet so we have something to work with.
    # This is not necessary if you are working with existing entities.
    inputTable = AtlasEntity(name="demo_sql_source", typeName="azure_sql_table",
                             qualified_name="pyapacheatlas://demo_update_lineagemapping_input", guid="-100")
    outputTable = AtlasEntity(name="demo_sql_target", typeName="azure_sql_table",
                              qualified_name="pyapacheatlas://demo_update_lineagemapping_output", guid="-101")
    # Upload these entities so we have something to work with
    # This will throw and exception if something goes wrong, otherwise
    # throw out the resulting json.
    _ = client.upload_entities([inputTable, outputTable])

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)
    # This is just a helper to fill in some demo data
    fill_in_workbook(file_path, excel_config)

    # ACTUAL WORK: We need a custom type that support the column mapping feature
    # We could have used a Data Factory pipeline but that is pretty complicated to set up
    typedefs = excel_reader.parse_entity_defs(file_path)
    _ = client.upload_typedefs(typedefs, force_update=True) # force_update to True so it's easier to repeat this step
    
    # This parses our excel file and creates a batch to upload
    lineage_with_mapping_processes = excel_reader.parse_update_lineage_with_mappings(file_path)

    # This is what is getting sent to your Atlas server
    print(json.dumps(lineage_with_mapping_processes,indent=2))

    results = client.upload_entities(lineage_with_mapping_processes)

    print(json.dumps(results, indent=2))

    print("Completed bulk upload of lineage processes successfully!")
    print("Search for 'demo_some_sql_process' to see your results.")
