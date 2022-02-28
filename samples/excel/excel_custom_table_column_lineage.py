import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient, AtlasEntity
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader


def fill_in_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    entityDef_sheet = wb[excel_config.entityDef_sheet]
    updateLineage_sheet = wb[excel_config.updateLineage_sheet]
    mapping_sheet = wb[excel_config.columnMapping_sheet]
    bulkEntity_sheet = wb[excel_config.bulkEntity_sheet]

    # BULK Sheet SCHEMA
    #"typeName", "name", "qualifiedName"
    # Adding a couple columns to show the power of this sheet
    # [Relationship] table, type
    entities_to_load = [
        ["hive_table", "hivetable01", "paa://hivetable01withcols",
            None, None],
        ["hive_column", "columnA", "paa://hivetable01withcols#colA",
            'paa://hivetable01withcols', 'string'],
        ["hive_column", "columnB", "paa://hivetable01withcols#colB",
            'paa://hivetable01withcols', 'long'],
        ["hive_column", "columnC", "paa://hivetable01withcols#colC",
            'paa://hivetable01withcols', 'int'],
        ["hive_table", "hivetable02withcols", "paa://hivetable02withcols",
            None, None],
        ["hive_column", "columnA", "paa://hivetable02withcols#colA",
            'paa://hivetable02withcols', 'string'],
        ["hive_column", "columnB", "paa://hivetable02withcols#colB",
            'paa://hivetable02withcols', 'long'],
        ["hive_column", "columnC", "paa://hivetable02withcols#colC",
            'paa://hivetable02withcols', 'int']
    ]

    # Need to adjust the default header to include our extra attributes
    bulkEntity_sheet['D1'] = '[Relationship] table'
    bulkEntity_sheet['E1'] = 'type'

    # Update Lineage Sheet SCHEMA
    # "Target typeName", "Target qualifiedName", "Source typeName",
    # "Source qualifiedName", "Process name", "Process qualifiedName",
    # "Process typeName"
    lineage_to_update = [
        ["hive_table", "paa://hivetable02withcols",
         "hive_table", "paa://hivetable01withcols",
         "custom_query",
         "paa://proc_update_lin_hive_tables",
         "customProcessWithMapping"
         ]
    ]

    # Mapping SCHEMA
    # "Source qualifiedName", "Source column", "Target qualifiedName", 
    # "Target column", "Process qualifiedName", "Process typeName",
    # "Process name"
    mapping_to_update = [
        ["paa://hivetable01withcols", "columnA", "paa://hivetable02withcols",
        "columnA", "paa://proc_update_lin_hive_tables", "customProcessWithMapping",
        "custom_query"
        ],
        ["paa://hivetable01withcols", "columnB", "paa://hivetable02withcols",
        "columnB", "paa://proc_update_lin_hive_tables", "customProcessWithMapping",
        "custom_query"
        ],
        ["paa://hivetable01withcols", "columnC", "paa://hivetable02withcols",
        "columnC", "paa://proc_update_lin_hive_tables", "customProcessWithMapping",
        "custom_query"
        ]        
    ]

    # EntityDef SCHEMA
    # "Entity TypeName", "name", "description",
    # "isOptional", "isUnique", "defaultValue",
    # "typeName", "displayName", "valuesMinCount",
    # "valuesMaxCount", "cardinality", "includeInNotification",
    # "indexType", "isIndexable", Entity superTypes (as an added non-default)
    entitydef_to_update = [
        [
        "customProcessWithMapping", "columnMapping", "stringified json to support mappings in Purview UI",
        True, None, None,
        "string", None, None,
        None, None, None,
        None, None,
        "Process"
        ]
    ]

    # Populate the excel template with samples above
    ## Bulk Entities
    entities_row_counter = 0
    for row in bulkEntity_sheet.iter_rows(min_row=2, max_col=5,
                                          max_row=len(entities_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = entities_to_load[entities_row_counter][idx]
        entities_row_counter += 1

    ## Update Lineage
    table_row_counter = 0
    for row in updateLineage_sheet.iter_rows(min_row=2, max_col=7,
                                             max_row=len(lineage_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = lineage_to_update[table_row_counter][idx]
        table_row_counter += 1
    
    # Update Column Mapping Sheet
    mapping_row_counter = 0
    for row in mapping_sheet.iter_rows(min_row=2, max_col=7,
                                             max_row=len(mapping_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = mapping_to_update[mapping_row_counter][idx]
        mapping_row_counter += 1

    # Update Entity Def Sheet
    entity_def_counter = 0
    for row in entityDef_sheet.iter_rows(min_row=2, max_col=15,
                                             max_row=len(entitydef_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = entitydef_to_update[entity_def_counter][idx]
        entity_def_counter += 1
    
    entityDef_sheet['O1'].value = 'Entity superTypes'

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file,
    generating new table and column entities, creating a custom type that
    supports Azure Purview's column mapping feature, and creating a
    custom Process that provides lineage and column mapping between the
    created entities.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", ""),
        authentication=oauth
    )

    # SETUP: This is just setting up the excel file for you
    file_path = "./demo_custom_table_column_lineage.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)
    # This is just a helper to fill in some demo data
    fill_in_workbook(file_path, excel_config)

    # ACTUAL WORK:
    # Parse your custom type def
    typedefs = excel_reader.parse_entity_defs(file_path)
    # force_update to True so it's easier to repeat this step
    _ = client.upload_typedefs(typedefs, force_update=True)
    
    # First extract the
    tables_cols = excel_reader.parse_bulk_entities(file_path)
    table_col_results = client.upload_entities(tables_cols)

    # This parses our excel file and creates a batch to upload
    lineage_with_mapping_processes = excel_reader.parse_update_lineage_with_mappings(file_path)

    # This is what is getting sent to your Atlas server
    lineage_results = client.upload_entities(lineage_with_mapping_processes)

    print(json.dumps([table_col_results, lineage_results], indent=2))

    print("Search for 'hivetable01withcols' to see your results.")
