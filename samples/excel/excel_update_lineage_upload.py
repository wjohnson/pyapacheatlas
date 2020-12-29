import json
import os

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import AtlasClient, AtlasEntity
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader


def fill_in_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    updateLineage_sheet = wb[excel_config.updateLineage_sheet]

    # Update Lineage Sheet SCHEMA
    # "Target typeName", "Target qualifiedName", "Source typeName",
    # "Source qualifiedName", "Process name", "Process qualifiedName",
    # "Process typeName"
    lineage_to_update = [
        ["hive_table", "pyapacheatlas://demo_update_lineage_output",
         "hive_table", "pyapacheatlas://demo_update_lineage_input",
         "demo_some_custom_hive_query",
         "pyapacheatlas://process_update_lineage_custom_hive_query",
         "Process"
         ]
    ]

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in updateLineage_sheet.iter_rows(min_row=2, max_col=7,
                                             max_row=len(lineage_to_update) + 1):
        for idx, cell in enumerate(row):
            cell.value = lineage_to_update[table_row_counter][idx]
        table_row_counter += 1

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file and
    generating a set of entities that would create or update a Process entity
    that links (applies lineage to) an input and output entity that already
    exists in your data catalog.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = AtlasClient(
        endpoint_url=os.environ.get("ENDPOINT_URL", ""),
        authentication=oauth
    )

    # SETUP: This is just setting up the excel file for you
    file_path = "./demo_update_lineage_upload.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # We are going to cheat here and create some entities before
    # we get to parsing the spreadsheet so we have something to work with.
    # This is not necessary if you are working with existing entities.
    inputTable = AtlasEntity(name="demo_hive_source", typeName="hive_table",
                             qualified_name="pyapacheatlas://demo_update_lineage_input", guid=-100)
    outputTable = AtlasEntity(name="demo_hive_target", typeName="hive_table",
                              qualified_name="pyapacheatlas://demo_update_lineage_output", guid=-101)
    # Upload these entities so we have something to work with
    # This will throw and exception if something goes wrong, otherwise
    # throw out the resulting json.
    _ = client.upload_entities([inputTable.to_json(), outputTable.to_json()])

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)
    # This is just a helper to fill in some demo data
    fill_in_workbook(file_path, excel_config)

    # ACTUAL WORK: This parses our excel file and creates a batch to upload
    lineage_processes = excel_reader.parse_update_lineage(file_path)

    # This is what is getting sent to your Atlas server
    # print(json.dumps(lineage_processes,indent=2))

    results = client.upload_entities(lineage_processes)

    print(json.dumps(results, indent=2))

    print("Completed bulk upload of lineage processes successfully!")
    print("Search for 'demo_some_custom_hive_query' to see your results.")
