import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient  # Communicate with your Atlas server
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader

from pyapacheatlas.core import TypeCategory


def fill_in_type_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    entityDef_sheet = wb[excel_config.entityDef_sheet]

    # ENTITYDEF Sheet SCHEMA
    # "Entity TypeName", "name", "description",
    # "isOptional", "isUnique", "defaultValue",
    # "typeName", "displayName", "valuesMinCount",
    # "valuesMaxCount", "cardinality", "includeInNotification",
    # "indexType", "isIndexable"
    attributes_to_load = [
        ["pyapacheatlas_custom_type", "fizz", "This will be the optional fizz attribute",
         None, None, None,
         None, None, None,
         None, None, None,
         None, None
         ],
        ["pyapacheatlas_custom_type", "buzz", "This will be the REQUIRED buzz attribute",
         False, None, None,
         None, None, None,
         None, None, None,
         None, None
         ],
    ]

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in entityDef_sheet.iter_rows(min_row=2, max_col=6,
                                         max_row=len(attributes_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = attributes_to_load[table_row_counter][idx]
        table_row_counter += 1

    wb.save(file_path)


def fill_in_entity_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    bulkEntity_sheet = wb[excel_config.bulkEntity_sheet]

    # BULK Sheet SCHEMA
    #"typeName", "name", "qualifiedName", "classifications"
    # Adding a couple columns to show the power of this sheet
    # fizz, buzz
    entities_to_load = [
        ["pyapacheatlas_custom_type", "custom_type_entity",
         "pyapacheatlas://example_from_custom_type", None,
         "abc", "123"
         ],
    ]

    # Need to adjust the default header to include our extra attributes
    bulkEntity_sheet['E1'] = 'fizz'
    bulkEntity_sheet['F1'] = 'buzz'

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in bulkEntity_sheet.iter_rows(min_row=2, max_col=6,
                                          max_row=len(entities_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = entities_to_load[table_row_counter][idx]
        table_row_counter += 1

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file,
    creating a custom type and then uploading an entity of that custom type.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", ""),
        authentication=oauth
    )

    # SETUP: This is just setting up the excel file for you
    file_path = "./demo_custom_type_and_entity_upload.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)
    # This is just a helper to fill in some demo data
    fill_in_type_workbook(file_path, excel_config)
    fill_in_entity_workbook(file_path, excel_config)

    # ACTUAL WORK: This parses our excel file and creates a batch to upload
    typedefs = excel_reader.parse_entity_defs(file_path)
    entities = excel_reader.parse_bulk_entities(file_path)

    # This is what is getting sent to your Atlas server
    # print(json.dumps(typedefs,indent=2))
    # print(json.dumps(entities,indent=2))

    type_results = client.upload_typedefs(typedefs, force_update=True)
    entity_results = client.upload_entities(entities)

    print(json.dumps(type_results, indent=2))
    print("\n")
    print(json.dumps(entity_results, indent=2))

    print("Completed type and bulk upload successfully!\nSearch for exampledataset to see your results.")
