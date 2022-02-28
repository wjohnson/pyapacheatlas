#########################
#  DEPRECATION WARNING  
# This sample uses deprecated features.
# Consider using excel_custom_table_column_lineage.py instead.
#########################

import json
import os

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient  # Communicate with your Atlas server
from pyapacheatlas.scaffolding import column_lineage_scaffold  # Create dummy types
# Read in the populated excel file.
# Customize header prefixes (e.g. "Sink" rather than "Target") and sheet names
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader
from pyapacheatlas.core.whatif import WhatIfValidator  # To do what if analysis


def fill_in_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    table_sheet = wb[excel_config.table_sheet]
    columns_sheet = wb[excel_config.column_sheet]

    # TABLE Sheet SCHEMA
    # "Target Table", "Target Type", "Target Classifications",
    # "Source Table", "Source Type", "Source Classifications",
    # "Process Name", "Process Type"
    # LIMITATION: Does not support multiple outputs from same process
    tables_to_load = [
        ["DestTable01", "demo_table", None, "SourceTable01",
            "demo_table", None, "Daily_ETL", "demo_process"],
        ["DestTable01", "demo_table", None, "SourceTable02",
            "demo_table", None, "Daily_ETL", "demo_process"],
        ["DestTable02", "demo_table", None, "SourceTable03",
            "demo_table", None, "Weekly_ETL", "demo_process"],
        ["DestTable03", "demo_table", None, None, None,
            None, "Stored_Proc:Do_Something", "demo_process"]
    ]
    # COLUMNS Sheet SCHEMA
    # "Target Table", "Target Column", "Target Classifications",
    # "Source Table", "Source Column", "Source Classifications",
    # "Transformation"
    columns_to_load = [
        ["DestTable01", "dest_c01", None, "SourceTable01", "source_c01",
         None, None],
        ["DestTable01", "dest_c02", None, "SourceTable01", "source_c02",
         None, None],
        # Demonstrate the ability to merge multiple columns
        ["DestTable01", "dest_combo01", None, "SourceTable01",
            "source_c03", None, "source_c03 + source_c04"],
        ["DestTable01", "dest_combo01", None, "SourceTable02",
            "source_c04", None, "source_c03 + source_c04"],
        # Demonstrate a simple, straightforward table with classifications
        ["DestTable02", "dest_c03", None, "SourceTable03",
            "source_c05", "MICROSOFT.PERSONAL.IPADDRESS", None],
        ["DestTable02", "dest_c04_express", None,
            None, None, None, "CURRENT_TIMESTAMP()"],
        # Demonstrate a table with no sources at all
        ["DestTable03", "dest_c100_express", None,
            None, None, None, "CURRENT_TIMESTAMP()"],
        ["DestTable03", "dest_c101_express",
            None, None, None, None, "RAND(100)"],
        ["DestTable03", "dest_c102_notransform", None, None, None,
         None, None],
    ]

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in table_sheet.iter_rows(min_row=2, max_col=8,
                                     max_row=len(tables_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = tables_to_load[table_row_counter][idx]
        table_row_counter += 1

    column_row_counter = 0
    for row in columns_sheet.iter_rows(min_row=2, max_col=7,
                                       max_row=len(columns_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = columns_to_load[column_row_counter][idx]
        column_row_counter += 1

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file,
    generating a table and column lineage set of entities, and then
    uploading the entities to your data catalog.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", ""),
        authentication=oauth
    )

    # Create an empty excel template to be populated
    file_path = "./atlas_excel_template.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    excel_reader.make_template(file_path, include_deprecated=True)

    fill_in_workbook(file_path, excel_config)

    # Generate the base atlas type defs for the demo of table and column lineage
    atlas_type_defs = column_lineage_scaffold(
        "demo", use_column_mapping=True,
        column_attributes=[{
            "name": "datatype",
            "typeName": "string",
            "isOptional": True,
            "cardinality": "SINGLE",
            "valuesMinCount": 1,
            "valuesMaxCount": 1,
            "isUnique": False,
            "isIndexable": False,
            "includeInNotification": False
        }]
    )
    # Alternatively, you can get all atlas types via...
    # atlas_type_defs = client.get_all_typedefs()

    input(">>>>Ready to upload type definitions?")
    # Upload scaffolded type defs and view the results of upload
    _upload_typedef = client.upload_typedefs(
        atlas_type_defs,
        force_update=True
    )
    print(json.dumps(_upload_typedef, indent=2))

    input(">>>>Review the above results to see what was uploaded.")

    # Generate the atlas entities!

    excel_results = excel_reader.parse_table_finegrain_column_lineages(
        file_path,
        atlas_type_defs,
        use_column_mapping=True
    )

    print("Results from excel transformation")
    print(json.dumps(excel_results, indent=2))

    input(">>>>Review the above results to see what your excel file contained")

    # Validate What IF
    whatif = WhatIfValidator(type_defs=atlas_type_defs)

    report = whatif.validate_entities(excel_results)

    if report["total"] > 0:
        print("There were errors in the provided typedefs")
        print(report)
        exit(1)
    else:
        print("There were no errors in the excel file")

    input(">>>>Review the what-if validation results above and get ready to upload your entities!")

    # Upload excel file's content to Atlas and view the guid assignments to confirm successful upload
    uploaded_entities = client.upload_entities(excel_results)
    print(json.dumps(uploaded_entities, indent=2))

    print("Completed uploads of demo!")
    # Be sure to clean up the excel file stored in file_path
